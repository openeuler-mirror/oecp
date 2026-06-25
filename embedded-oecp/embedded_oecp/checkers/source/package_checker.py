# **********************************************************************************
# Copyright (c) Huawei Technologies Co., Ltd. 2026-2026. All rights reserved.
# [embedded-oecp] is licensed under the Mulan PSL v2.
# You can use this software according to the terms and conditions of the Mulan PSL v2.
# You may obtain a copy of Mulan PSL v2 at:
#     http://license.coscl.org.cn/MulanPSL2
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY OR FIT FOR A PARTICULAR
# PURPOSE.
# See the Mulan PSL v2 for more details.
# Author: lixinyu
# Create: 2025-01-01
# Description: embedded-oecp utility
# **********************************************************************************
import os
import re
import subprocess
from typing import List, Dict, Optional
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from embedded_oecp.checkers import BaseChecker
from embedded_oecp.models import TestResult
from embedded_oecp.utils.shell import run_command
from embedded_oecp.utils.logger import get_logger

OPENEULER_SRC_HOSTS = [
    "atomgit.com/src-openeuler",
    "gitee.com/src-openeuler",
    "atomgit.com/openeuler",
    "gitee.com/openeuler",
]

COMMUNITY_RATIO_THRESHOLD = 0.70

SKIP_SUFFIXES = ("-native", "-cross", "-crosssdk", "-initial", "-canonical", "-nativesdk")


class PackageChecker(BaseChecker):
    def __init__(self, config: dict, work_dir: str):
        super().__init__(config, work_dir)
        self._evidence_dir = ""

    def run(self) -> List[TestResult]:
        self._evidence_dir = os.path.join(
            self.config.get("output_dir", "/tmp/embedded-oecp/report"), "evidence",
        )
        os.makedirs(self._evidence_dir, exist_ok=True)
        results = []
        results.append(self._check_and_export_packages())
        return results

    def _get_build_dir(self) -> str:
        d = self.get_build_dir()
        if d and os.path.isdir(d):
            return d
        return ""

    def _get_src_dir(self, build_dir: str) -> str:
        return self.get_source_dir()

    def _parse_buildlist(self, build_dir: str) -> List[str]:
        bl_path = os.path.join(build_dir, "pn-buildlist")
        if not os.path.isfile(bl_path):
            return []
        packages = []
        with open(bl_path, "r") as f:
            for line in f:
                name = line.strip()
                if name:
                    packages.append(name)
        return packages

    def _should_skip(self, pkg_name: str) -> bool:
        return any(pkg_name.endswith(s) for s in SKIP_SUFFIXES)

    def _query_local_name(self, pkg_name: str, build_dir: str) -> Optional[str]:
        cmd = ["oebuild", "bitbake-getvar", "-r", pkg_name, "--value", "OPENEULER_LOCAL_NAME"]
        try:
            proc = subprocess.run(
                cmd, cwd=build_dir, capture_output=True, text=True, timeout=30,
            )
            lines = [line.strip() for line in proc.stdout.strip().splitlines() if line.strip()]
            if not lines:
                return None
            last_line = lines[-1]
            parts = last_line.split(" - ", 2)
            value = parts[-1].strip() if len(parts) >= 3 else last_line.strip()
            if value == "None" or not value:
                return None
            return value
        except Exception:
            return None

    def _serial_query_local_names(self, pkg_names: List[str], build_dir: str) -> Dict[str, Optional[str]]:
        logger = get_logger()
        results = {}
        total = len(pkg_names)
        for i, name in enumerate(pkg_names):
            local_name = self._query_local_name(name, build_dir)
            results[name] = local_name
            if local_name:
                logger.info(f"[{i+1}/{total}] {name} -> {local_name}")
            elif (i + 1) % 20 == 0:
                logger.info(f"[{i+1}/{total}] querying...")
        found = sum(1 for v in results.values() if v)
        logger.info(f"bitbake-getvar queried {total} packages, {found} have OPENEULER_LOCAL_NAME")
        return results

    def _get_repo_remote_url(self, repo_dir: str) -> str:
        stdout, _, rc = run_command("git remote", cwd=repo_dir, timeout=10)
        if rc != 0:
            return ""
        for remote in stdout.strip().split("\n"):
            remote = remote.strip()
            if not remote:
                continue
            stdout2, _, rc2 = run_command(f"git remote get-url {remote}", cwd=repo_dir, timeout=10)
            if rc2 == 0:
                return stdout2.strip()
        return ""

    def _get_repo_head_commit(self, repo_dir: str) -> str:
        stdout, _, rc = run_command("git rev-parse HEAD", cwd=repo_dir, timeout=10)
        return stdout.strip() if rc == 0 else ""

    def _load_manifest(self, build_dir: str) -> Dict[str, dict]:
        logger = get_logger()

        baseline_repo = self.config.get("baseline_repo", "")
        baseline_branch = self.config.get("baseline_branch", "")
        if baseline_repo and baseline_branch:
            manifest_url = self._build_manifest_url()
            if manifest_url:
                manifest = self._fetch_remote_manifest(manifest_url)
                if manifest:
                    return manifest

        src_dir = self._get_src_dir(build_dir)
        candidates = []
        if src_dir:
            candidates.append(os.path.join(src_dir, "yocto-meta-openeuler", ".oebuild", "manifest.yaml"))
        if build_dir:
            candidates.append(os.path.normpath(
                os.path.join(build_dir, "..", "..", "src", "yocto-meta-openeuler", ".oebuild", "manifest.yaml")
            ))
        for p in candidates:
            if os.path.isfile(p):
                logger.info(f"Loading manifest from local: {p}")
                with open(p, "r", encoding="utf-8") as f:
                    data = yaml.safe_load(f) or {}
                return data.get("manifest_list", {})
        logger.warning("No manifest found (remote or local)")
        return {}

    def _fetch_remote_manifest(self, url: str) -> Dict[str, dict]:
        logger = get_logger()
        logger.info(f"Fetching manifest from {url}")
        try:
            import urllib.request
            req = urllib.request.Request(url, headers={"User-Agent": "embedded-oecp/1.0"})
            with urllib.request.urlopen(req, timeout=30) as resp:
                content = resp.read().decode("utf-8")
            data = yaml.safe_load(content) or {}
            manifest = data.get("manifest_list", {})
            if manifest:
                logger.info(f"Remote manifest loaded: {len(manifest)} entries ({len(content)} bytes)")
                return manifest
            logger.warning("Remote manifest has no manifest_list")
        except Exception as e:
            logger.warning(f"Failed to fetch remote manifest: {e}")
        return {}

    def _build_manifest_url(self) -> str:
        baseline_repo = self.config.get("baseline_repo", "")
        baseline_branch = self.config.get("baseline_branch", "")
        m = re.match(r'https?://[^/]+/([^/]+/[^/]+?)(?:\.git)?$', baseline_repo)
        if not m:
            return ""
        return f"https://raw.atomgit.com/{m.group(1)}/raw/{baseline_branch}/.oebuild/manifest.yaml"

    def _is_community_host(self, url: str) -> bool:
        return any(host in url.lower() for host in OPENEULER_SRC_HOSTS)

    def _check_and_export_packages(self) -> TestResult:
        logger = get_logger()
        result = TestResult(
            category="源码", sub_item="其他软件包",
            requirement="镜像依赖包来源检查，社区包占比≥70%，commit ID 与基线一致",
            acceptance_criteria="生成镜像依赖包 Excel 列表，社区包占比≥70%",
        )

        build_dir = self._get_build_dir()
        if not build_dir:
            result.fail_test(detail="未配置镜像编译目录", remediation="请使用 -d 指定编译目录")
            return result

        src_dir = self._get_src_dir(build_dir)
        if not src_dir:
            result.fail_test(detail="未找到 src 目录")
            return result

        openeuler_dir = os.path.join(src_dir, "openeuler")
        if not os.path.isdir(openeuler_dir):
            result.fail_test(detail=f"未找到 {openeuler_dir} 目录")
            return result

        manifest = self._load_manifest(build_dir)
        manifest_url = self._build_manifest_url()

        buildlist = self._parse_buildlist(build_dir)
        if not buildlist:
            logger.info("pn-buildlist 不存在，尝试生成...")
            run_command("oebuild bitbake -g openeuler-image", cwd=build_dir, timeout=600)
            buildlist = self._parse_buildlist(build_dir)
            if not buildlist:
                result.fail_test(detail="无法获取 pn-buildlist")
                return result

        query_list = [p for p in buildlist if not self._should_skip(p)]
        skipped = len(buildlist) - len(query_list)
        logger.info(
            f"pn-buildlist: {len(buildlist)} total, "
            f"{len(query_list)} to query (skipped {skipped} native/cross)"
        )

        local_name_map = self._serial_query_local_names(query_list, build_dir)

        pkg_records = []
        for pkg_name in query_list:
            local_name = local_name_map.get(pkg_name)
            if not local_name:
                continue

            pkg_dir = os.path.join(openeuler_dir, local_name)
            if not os.path.isdir(pkg_dir) or not os.path.isdir(os.path.join(pkg_dir, ".git")):
                continue

            remote_url = self._get_repo_remote_url(pkg_dir)
            commit_hash = self._get_repo_head_commit(pkg_dir)

            url_is_community = self._is_community_host(remote_url) if remote_url else False
            baseline_match = False
            baseline_info = None

            if manifest:
                entry = manifest.get(local_name)
                if entry and isinstance(entry, dict):
                    baseline_url = entry.get("remote_url", "")
                    baseline_version = entry.get("version", "")
                    baseline_info = {"remote_url": baseline_url, "version": baseline_version}
                    if commit_hash and commit_hash == baseline_version:
                        baseline_match = True

            is_community = url_is_community and baseline_match

            pkg_records.append({
                "name": local_name,
                "remote_url": remote_url or "N/A",
                "commit_hash": commit_hash or "N/A",
                "is_community": is_community,
                "url_is_community": url_is_community,
                "baseline_match": baseline_match,
                "baseline_info": baseline_info,
            })

        logger.info(f"Resolved packages with git repos: {len(pkg_records)}")

        community_count = sum(1 for r in pkg_records if r["is_community"])
        baseline_matched = sum(1 for r in pkg_records if r["baseline_match"])
        url_community_count = sum(1 for r in pkg_records if r["url_is_community"])
        total = len(pkg_records)
        ratio = community_count / total if total > 0 else 0
        pct = ratio * 100

        excel_path = self._generate_excel(pkg_records, manifest, manifest_url)

        if ratio >= COMMUNITY_RATIO_THRESHOLD:
            result.pass_test(
                detail=(
                    f"pn-buildlist 共 {len(buildlist)} 个包，"
                    f"查询 {len(query_list)} 个（跳过 {skipped} 个 native/cross），"
                    f"解析到 {total} 个源码包，"
                    f"社区包 {community_count} 个({pct:.1f}%)，"
                    f"基线匹配 {baseline_matched} 个。"
                    f"社区包占比 {pct:.1f}% >= 70%，符合要求。"
                ),
                evidence=f"excel: {excel_path}",
            )
        else:
            result.fail_test(
                detail=(
                    f"pn-buildlist 共 {len(buildlist)} 个包，"
                    f"查询 {len(query_list)} 个（跳过 {skipped} 个 native/cross），"
                    f"解析到 {total} 个源码包，"
                    f"社区包 {community_count} 个({pct:.1f}%)，"
                    f"基线匹配 {baseline_matched} 个。"
                    f"社区包占比 {pct:.1f}% < 70%，不符合要求。"
                ),
                evidence=f"excel: {excel_path}",
                remediation="请增加来自 openEuler 社区仓的软件包",
            )

        return result

    def _generate_excel(self, pkg_records: List[dict], manifest: dict, manifest_url: str) -> str:
        wb = Workbook()

        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        link_font = Font(color="0563C1", underline="single", size=11)

        ws2 = wb.active
        ws2.title = "社区基线(Manifest)"

        ws2.merge_cells("A1:C1")
        src_cell = ws2["A1"]
        src_cell.value = f"数据来源: {manifest_url}" if manifest_url else "数据来源: manifest.yaml（本地）"
        src_cell.font = Font(bold=True, size=11, color="333333")

        headers2 = ["包名", "源码仓地址", "版本Hash"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=2, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        manifest_sorted = sorted(manifest.items()) if manifest else []
        for row_idx, (name, info) in enumerate(manifest_sorted, 3):
            if not isinstance(info, dict):
                continue
            ws2.cell(row=row_idx, column=1, value=name).border = thin_border
            ws2.cell(row=row_idx, column=2, value=info.get("remote_url", "")).border = thin_border
            ws2.cell(row=row_idx, column=3, value=info.get("version", "")).border = thin_border

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 60
        ws2.column_dimensions["C"].width = 45

        manifest_name_to_row = {}
        for row_idx, (name, info) in enumerate(manifest_sorted, 3):
            manifest_name_to_row[name] = row_idx

        ws1 = wb.create_sheet(title="镜像依赖包列表", index=0)

        headers1 = ["包名", "源码仓地址", "版本Hash", "是否社区包", "是否基线匹配", "基线对应行"]
        for col, h in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, rec in enumerate(pkg_records, 2):
            ws1.cell(row=row_idx, column=1, value=rec["name"]).border = thin_border
            ws1.cell(row=row_idx, column=2, value=rec["remote_url"]).border = thin_border
            ws1.cell(row=row_idx, column=3, value=rec["commit_hash"]).border = thin_border

            is_comm = "是" if rec["is_community"] else "否"
            cell_comm = ws1.cell(row=row_idx, column=4, value=is_comm)
            cell_comm.border = thin_border
            cell_comm.alignment = Alignment(horizontal="center")
            if rec["is_community"]:
                cell_comm.fill = green_fill

            cell_match = ws1.cell(row=row_idx, column=5)
            cell_match.border = thin_border
            cell_match.alignment = Alignment(horizontal="center")
            if rec["url_is_community"]:
                match_val = "是" if rec["baseline_match"] else "否"
                cell_match.value = match_val
                if rec["baseline_match"]:
                    cell_match.fill = green_fill
                else:
                    cell_match.fill = red_fill
            else:
                cell_match.value = "N/A"

            cell_link = ws1.cell(row=row_idx, column=6)
            cell_link.border = thin_border
            cell_link.alignment = Alignment(horizontal="center")

            baseline_row = manifest_name_to_row.get(rec["name"])
            if baseline_row:
                link_target = f"'社区基线(Manifest)'!A{baseline_row}"
                cell_link.value = f"基线第{baseline_row}行"
                cell_link.hyperlink = link_target
                cell_link.font = link_font
            elif rec["url_is_community"]:
                cell_link.value = "基线中无记录"
                cell_link.font = Font(color="999999", size=11)

        ws1.column_dimensions["A"].width = 30
        ws1.column_dimensions["B"].width = 60
        ws1.column_dimensions["C"].width = 45
        ws1.column_dimensions["D"].width = 14
        ws1.column_dimensions["E"].width = 14
        ws1.column_dimensions["F"].width = 16

        excel_path = os.path.join(self._evidence_dir, "package_dependency_list.xlsx")
        wb.save(excel_path)
        logger = get_logger()
        logger.info(f"Excel saved: {excel_path}")
        return excel_path
