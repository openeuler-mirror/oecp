# **********************************************************************************
# Copyright (c) Huawei Technologies Co., Ltd. 2026-2026. All rights reserved.
# [embedded-oecp] is licensed under the Mulan PSL v2.
# You can use this software according to the terms and conditions of the Mulan PSL v2.
# You may obtain a copy of Mulan PSL v2 at:
#     http://license.coscl.org.cn/MulanPSL2
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY OR FIT FOR A PARTICULAR
# PURPOSE.
# See the Mulan PSL v2 for more details.
# Author: lixinyu
# Create: 2025-01-01
# Description: embedded-oecp utility
# **********************************************************************************
import os
import csv
import subprocess
from typing import List, Dict, Optional

import yaml

from embedded_oecp.checkers import BaseChecker
from embedded_oecp.models import TestResult
from embedded_oecp.utils.logger import get_logger
from embedded_oecp.utils.terminal_screenshot import render_terminal_screenshot

PKG_CONSISTENCY_THRESHOLD = 0.70

ARCHIVE_EXTS = ('.tar.', '.tgz', '.tbz', '.tbz2', '.zip', '.gz', '.xz', '.bz2', '.tar')


class PkgListChecker(BaseChecker):
    def __init__(self, config: dict, work_dir: str):
        super().__init__(config, work_dir)
        self._evidence_dir = ""
        self._src_uri_cache: Dict[str, str] = {}

    def run(self) -> List[TestResult]:
        self._evidence_dir = os.path.join(
            self.config.get("output_dir", "/tmp/embedded-oecp/report"), "evidence",
        )
        os.makedirs(self._evidence_dir, exist_ok=True)
        return [self._check_pkglist_consistency()]

    def _get_build_dir(self) -> str:
        return self.get_build_dir()

    def _get_machine(self, build_dir: str) -> str:
        deploy_dir = os.path.join(build_dir, "tmp", "deploy", "images")
        if os.path.isdir(deploy_dir):
            for d in os.listdir(deploy_dir):
                if os.path.isdir(os.path.join(deploy_dir, d)):
                    return d
        return ""

    def _find_manifest_file(self, build_dir: str, machine: str) -> Optional[str]:
        deploy_dir = os.path.join(build_dir, "tmp", "deploy", "images", machine)
        if not os.path.isdir(deploy_dir):
            return None
        for f in sorted(os.listdir(deploy_dir), reverse=True):
            if f.endswith(".manifest"):
                return os.path.join(deploy_dir, f)
        return None

    def _parse_manifest(self, manifest_path: str) -> List[Dict]:
        pkgs = []
        with open(manifest_path, "r") as f:
            for line in f:
                parts = line.strip().split()
                if len(parts) >= 3:
                    pkgs.append({"name": parts[0], "arch": parts[1], "version": parts[2]})
                elif len(parts) == 2:
                    pkgs.append({"name": parts[0], "arch": "", "version": parts[1]})
        return pkgs

    def _load_pkgdata_map(self, build_dir: str, machine: str) -> Dict[str, dict]:
        logger = get_logger()
        runtime_dir = os.path.join(build_dir, "tmp", "pkgdata", machine, "runtime")
        if not os.path.isdir(runtime_dir):
            logger.warning(f"pkgdata runtime dir not found: {runtime_dir}")
            return {}

        pkg_map: Dict[str, dict] = {}
        for fname in os.listdir(runtime_dir):
            fpath = os.path.join(runtime_dir, fname)
            if not os.path.isfile(fpath):
                continue
            local_name = ""
            pn = ""
            pv = ""
            renamed = {}
            try:
                with open(fpath, "r") as f:
                    for line in f:
                        if line.startswith("OPENEULER_LOCAL_NAME:"):
                            local_name = line.split(":", 1)[1].strip()
                        elif line.startswith("PN:"):
                            pn = line.split(":", 1)[1].strip()
                        elif line.startswith("PV:"):
                            pv = line.split(":", 1)[1].strip()
                        elif line.startswith("PKG:"):
                            parts = line.split(":", 2)
                            if len(parts) == 3:
                                renamed[parts[1]] = parts[2].strip()
            except Exception as e:
                logger.debug(f"Skipped pkgdata file {fname}: {e}")
                continue
            if local_name:
                info = {"local_name": local_name, "pn": pn, "pv": pv}
                pkg_map[fname] = info
                for pn_key, binary_name in renamed.items():
                    if binary_name and binary_name != pn_key and binary_name not in pkg_map:
                        pkg_map[binary_name] = info
        logger.info(f"Loaded pkgdata map: {len(pkg_map)} entries")
        return pkg_map

    def _get_source_remote_version(self, local_name: str) -> Optional[Dict]:
        source_dir = self.get_source_dir()
        if not source_dir:
            return None

        pkg_dir = os.path.join(source_dir, "openeuler", local_name)
        if not os.path.isdir(pkg_dir):
            return None

        remote = ""
        for remote_name in ("upstream", "origin"):
            try:
                result = subprocess.run(
                    ["git", "remote", "get-url", remote_name],
                    capture_output=True, text=True, cwd=pkg_dir, timeout=10,
                )
                if result.returncode == 0 and result.stdout.strip():
                    remote = result.stdout.strip()
                    break
            except Exception as e:
                logger.debug(f"Failed to get git remote for {local_name}: {e}")

        version = ""
        spec_files = [f for f in os.listdir(pkg_dir) if f.endswith(".spec")]
        if spec_files:
            try:
                with open(os.path.join(pkg_dir, spec_files[0]), "r") as f:
                    for line in f:
                        sline = line.strip()
                        if sline.startswith("Version:"):
                            version = sline.split(":", 1)[1].strip()
                            break
            except Exception as e:
                logger.debug(f"Failed to read spec file for {local_name}: {e}")

        if not remote:
            return None
        return {"remote": remote, "version": version}

    def _build_manifest_raw_url(self, repo: str, branch: str) -> str:
        import re
        repo = repo.rstrip("/")
        if repo.endswith(".git"):
            repo = repo[:-4]
        m = re.match(r'https?://atomgit\.com/([^/]+)/(.+)', repo)
        if m:
            org, name = m.group(1), m.group(2)
            return f"https://raw.atomgit.com/{org}/{name}/raw/{branch}/.oebuild/manifest.yaml"
        return ""

    def _load_manifest_baseline(self) -> Dict:
        logger = get_logger()
        source_dir = self.get_source_dir()
        if source_dir:
            manifest_path = os.path.join(source_dir, "yocto-meta-openeuler", ".oebuild", "manifest.yaml")
            if os.path.isfile(manifest_path):
                logger.info(f"Loading manifest baseline: {manifest_path}")
                try:
                    with open(manifest_path, "r") as f:
                        data = yaml.safe_load(f) or {}
                    manifest_list = data.get("manifest_list", {})
                    if isinstance(manifest_list, dict):
                        logger.info(f"Baseline loaded: {len(manifest_list)} entries")
                        return manifest_list
                except Exception as e:
                    logger.warning(f"Failed: {e}")

        repo = self.config.get("baseline_repo", "")
        branch = self.config.get("baseline_branch", "master")
        raw_url = self._build_manifest_raw_url(repo, branch)
        if raw_url:
            logger.info(f"Fetching remote manifest.yaml: {raw_url}")
            try:
                import urllib.request
                with urllib.request.urlopen(raw_url, timeout=30) as resp:
                    data = yaml.safe_load(resp.read()) or {}
                manifest_list = data.get("manifest_list", {})
                if isinstance(manifest_list, dict):
                    logger.info(f"Remote baseline loaded: {len(manifest_list)} entries")
                    return manifest_list
            except Exception as e:
                logger.warning(f"Failed to fetch remote manifest: {e}")
        else:
            logger.warning(f"Cannot build manifest URL from baseline_repo={repo}")
        return {}

    def _is_openeuler_remote(self, remote: str) -> bool:
        rl = remote.lower()
        return "openeuler" in rl or "src-openeuler" in rl

    # ---- 辅助配方判断 ----

    def _get_bitbake_src_uri(self, pn: str) -> str:
        logger = get_logger()
        if pn in self._src_uri_cache:
            return self._src_uri_cache[pn]
        build_dir = self._get_build_dir()
        src_uri = ""
        try:
            result = subprocess.run(
                ["oebuild", "bitbake", "-e", pn],
                capture_output=True, text=True, timeout=60, cwd=build_dir,
            )
            for line in result.stdout.splitlines():
                if line.startswith("SRC_URI="):
                    src_uri = line.strip()
                    break
        except Exception as e:
            logger.warning(f"bitbake -e {pn} failed: {e}")
        self._src_uri_cache[pn] = src_uri
        logger.info(f"SRC_URI[{pn}]: {src_uri[:120]}")
        return src_uri

    def _is_auxiliary_recipe(self, pn: str, pv: str) -> bool:
        if pv != "1.0":
            return False
        src_uri_line = self._get_bitbake_src_uri(pn)
        if not src_uri_line or "=" not in src_uri_line:
            return True
        uri_val = src_uri_line.split("=", 1)[1].strip().strip('"')
        if not uri_val:
            return True
        entries = uri_val.split()
        for entry in entries:
            low = entry.lower()
            for ext in ARCHIVE_EXTS:
                if ext in low:
                    return False
        return True

    # ---- 主逻辑 ----

    def _check_pkglist_consistency(self) -> TestResult:
        logger = get_logger()
        result = TestResult(
            category="构建", sub_item="包列表",
            requirement="社区包版本一致性检查",
            acceptance_criteria="软件包与 openEuler 社区基线快照版本比对，记录到表格",
        )

        build_dir = self._get_build_dir()
        if not build_dir:
            result.fail_test(detail="未配置构建目录", remediation="请配置 image_build.dir")
            return result

        machine = self._get_machine(build_dir)
        if not machine:
            result.fail_test(detail="未找到 MACHINE 名称")
            return result

        manifest_path = self._find_manifest_file(build_dir, machine)
        if not manifest_path:
            result.fail_test(detail=f"未在 tmp/deploy/images/{machine} 下找到 manifest 文件")
            return result

        pkgs = self._parse_manifest(manifest_path)
        if not pkgs:
            result.fail_test(detail="manifest 文件为空")
            return result

        logger.info(f"Manifest: {os.path.basename(manifest_path)}, {len(pkgs)} packages")

        baseline = self._load_manifest_baseline()
        if not baseline:
            result.fail_test(detail="未加载到社区基线快照（manifest.yaml）")
            return result

        pkgdata_map = self._load_pkgdata_map(build_dir, machine)

        table_rows = []
        community_matched = 0
        community_unmatched = 0
        non_community = 0
        no_source = 0
        auxiliary_skipped = 0

        for pkg in pkgs:
            pkg_name = pkg["name"]
            pkg_version = pkg["version"]

            if pkg_name.startswith("packagegroup"):
                continue

            row = {
                "name": pkg_name,
                "version": pkg_version,
                "consistent": "",
                "community_remote": "",
                "community_version": "",
            }

            info = pkgdata_map.get(pkg_name)
            if info:
                local_name = info["local_name"]
                pn = info["pn"]
                pv = info["pv"]
            else:
                local_name = ""
                pn = pkg_name
                pv = ""

            base_entry = None
            if local_name:
                base_entry = baseline.get(local_name) or baseline.get(pkg_name)

            if base_entry:
                src_info = self._get_source_remote_version(local_name) if local_name else None
                if not src_info:
                    row["consistent"] = "社区包但未找到源码目录"
                    no_source += 1
                    table_rows.append(row)
                    continue

                remote = src_info["remote"]

                if isinstance(base_entry, dict):
                    base_remote = base_entry.get("remote_url", "")
                    base_version = base_entry.get("version", "")
                else:
                    base_remote = ""
                    base_version = str(base_entry) if base_entry else ""

                if remote == base_remote:
                    row["consistent"] = "与社区包版本一致"
                    row["community_remote"] = base_remote
                    row["community_version"] = base_version
                    community_matched += 1
                else:
                    row["consistent"] = "与社区包版本不一致"
                    row["community_remote"] = base_remote
                    row["community_version"] = base_version
                    community_unmatched += 1

                table_rows.append(row)
            else:
                if self._is_auxiliary_recipe(pn, pv):
                    auxiliary_skipped += 1
                    continue

                row["consistent"] = "非社区包"
                non_community += 1
                table_rows.append(row)

        self._save_excel(table_rows)

        total = len(table_rows)
        ratio = community_matched / total if total > 0 else 0
        pct = ratio * 100

        logger.info(f"Result: matched={community_matched}, unmatched={community_unmatched}, "
                     f"non_community={non_community}, no_source={no_source}, "
                     f"auxiliary_skipped={auxiliary_skipped}, total={total}")

        bar_len = 40
        filled = int(bar_len * ratio)
        bar = "█" * filled + "░" * (bar_len - filled)

        lines = [
            {"text": "# 社区包版本一致性检查", "color": "#6c7086"},
            {"text": f"# manifest: {os.path.basename(manifest_path)}", "color": "#6c7086"},
            {"text": "", "color": "#cdd6f4"},
            {"text": f"软件包总数（排除packagegroup/辅助配方）: {total}", "color": "#cdd6f4"},
            {"text": f"与社区包版本一致:   {community_matched}", "color": "#a6e3a1"},
            {"text": f"与社区包版本不一致: {community_unmatched}", "color": "#f38ba8"},
            {"text": f"非社区包:           {non_community}", "color": "#fab387"},
            {"text": f"社区包无源码目录:   {no_source}", "color": "#f9e2af"},
            {"text": f"辅助配方（已排除）: {auxiliary_skipped}", "color": "#6c7086"},
            {"text": "", "color": "#cdd6f4"},
            {"text": f"一致率: [{bar}] {pct:.1f}%", "type": "highlight", "color": "#f9e2af"},
            {"text": "", "color": "#cdd6f4"},
        ]

        if ratio >= PKG_CONSISTENCY_THRESHOLD:
            lines.append({"text": f"社区包版本一致率 {pct:.1f}% >= 70%，符合要求", "type": "success"})
        else:
            lines.append({"text": f"社区包版本一致率 {pct:.1f}% < 70%，不符合要求", "type": "error"})

        path = os.path.join(self._evidence_dir, "pkglist_step1_consistency.png")
        render_terminal_screenshot("Step 1: 社区包版本一致性检查", lines, path)

        if ratio >= PKG_CONSISTENCY_THRESHOLD:
            result.pass_test(
                detail=f"社区包版本一致率: {community_matched}/{total} = {pct:.1f}%（≥70%）",
                evidence=f"matched={community_matched}, unmatched={community_unmatched}, "
                          f"non_community={non_community}, no_source={no_source}, "
                          f"auxiliary_skipped={auxiliary_skipped}, screenshot={path}",
            )
        else:
            result.fail_test(
                detail=f"社区包版本一致率: {community_matched}/{total} = {pct:.1f}%（<70%）",
                evidence=f"screenshot={path}",
                remediation="请确保使用与社区基线一致的软件包版本",
            )
        return result

    def _save_excel(self, rows: List[Dict]):
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            self._save_csv(rows)
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "包列表检查结果"

        headers = ["软件包名", "版本", "是否与社区包版本一致", "社区包remote", "社区包version"]
        ws.append(headers)
        for col in range(1, 6):
            ws.cell(row=1, column=col).font = Font(bold=True)

        for row in rows:
            ws.append([
                row.get("name", ""),
                row.get("version", ""),
                row.get("consistent", ""),
                row.get("community_remote", ""),
                row.get("community_version", ""),
            ])

        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in r:
                if cell.value == "与社区包版本一致":
                    cell.font = Font(color="008000")
                elif cell.value and "不一致" in str(cell.value):
                    cell.font = Font(color="CC0000")
                elif cell.value in ("非社区包", "社区包但未找到源码目录"):
                    cell.font = Font(color="999999")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 45

        xlsx_path = os.path.join(self._evidence_dir, "package_dependency_list.xlsx")
        wb.save(xlsx_path)
        get_logger().info(f"Excel saved: {xlsx_path} ({len(rows)} rows)")

    def _save_csv(self, rows: List[Dict]):
        csv_path = os.path.join(self._evidence_dir, "package_dependency_list.csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["软件包名", "版本", "是否与社区包版本一致", "社区包remote", "社区包version"])
            for row in rows:
                writer.writerow([
                    row.get("name", ""), row.get("version", ""),
                    row.get("consistent", ""), row.get("community_remote", ""),
                    row.get("community_version", ""),
                ])
        get_logger().info(f"CSV saved: {csv_path}")
