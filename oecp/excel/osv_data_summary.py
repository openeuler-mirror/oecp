# -*- encoding=utf-8 -*-
"""
# **********************************************************************************
# Copyright (c) Huawei Technologies Co., Ltd. 2021-2021. All rights reserved.
# [oecp] is licensed under the Mulan PSL v2.
# You can use this software according to the terms and conditions of the Mulan PSL v2.
# You may obtain a copy of Mulan PSL v2 at:
#     http://license.coscl.org.cn/MulanPSL2
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY OR FIT FOR A PARTICULAR
# PURPOSE.
# See the Mulan PSL v2 for more details.
# **********************************************************************************
"""

import json
import shutil
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from pathlib import Path

from oecp.result.compress import gen_hash_key
from oecp.result.constants import *

logger = logging.getLogger("oecp")


class DataExcelFile:
    """
    Excel table display of the final result
    """

    def __init__(self):
        self.path = Path(Path(__file__).parents[0])
        self.excel_template = Path(Path(__file__).parents[1].joinpath("conf", "excel_template"))
        self.green = PatternFill("solid", fgColor="90EE90")
        self.red = PatternFill("solid", fgColor="FFE4E1")
        self.osv_summary_path = None
        self.benchmark_json = Path(self.excel_template).joinpath("benchmark.json")
        self.green_font = Font(size=FONT_SIZE, color="00FF00")
        self.red_font = Font(size=FONT_SIZE, color="FF0000")
        self.table = None
        self.sheet = None
        self.folder_path = None
        self.tools_result = {}
        self.conclusion = ''

    @staticmethod
    def convert_data(data):
        """[summary]

        Args:
            data ([type]): [description]

        Returns:
            [type]: [description]
        """
        return "%.2f%%" % (data * 100)

    @staticmethod
    def obtain_base_os(side_a):
        base_so = side_a.split("-")[:4]
        for part in base_so:
            if '.iso' in part or "aarch64" == part or "x86_64" == part:
                base_so.remove(part)
        return " ".join(base_so)

    def create_folder(self, root_path, osv_title):
        """
        The folder where the results are stored
        """
        self.folder_path = Path(Path(root_path).joinpath(osv_title))
        Path(self.folder_path).mkdir(parents=True, exist_ok=True)

    def copy_file(self, root_path, osv_title):
        """
        copy file
        """
        self.create_folder(root_path, osv_title)
        self.osv_summary_path = Path(self.folder_path).joinpath("osv_data_summary.xlsx")
        template_path = Path(self.excel_template).joinpath("template.xlsx")
        shutil.copy(template_path, self.osv_summary_path)

    def open_excel(self):
        """
        Operate the object of excel
        """
        self.table = load_workbook(self.osv_summary_path)
        self.sheet = self.table["Sheet1"]

    def close(self):
        """
        Close the object that operates excel
        """
        self.table.close()

    def read_excel_content(self):
        """
        Statistic results
        """
        self.open_excel()
        conclusion = "通过"
        for row_num in REQUIRED_ROW:
            result = self.sheet.cell(row=row_num, column=SEVEN_COLUMN).value
            if result == "NO PASS":
                conclusion = "不通过"
                break
        self.close()
        return conclusion

    def read_benchmark_json(self):
        """
        read benchmark json
        """

        with open(self.benchmark_json, 'r', encoding='utf8') as fp:
            benchmark_criteria = json.load(fp)
        return benchmark_criteria

    def result_comparison(self, similarity, content, option, data, index):
        """
        Calculate whether to pass according to the result
        Args:
            similarity: similarity
            content: content
            option: option
            data: data
            index: index

        Returns:
            None
        """
        for choice, benchmark in content.items():
            if choice == "L1 and L2 package":
                lo_pkg = similarity.get("level1 pkg")
                lt_pkg = similarity.get("level2 pkg")
                result = "PASS" if lo_pkg >= benchmark[0]["level1 pkg"]["standard"] and lt_pkg >= \
                                   benchmark[1]["level2 pkg"]["standard"] else "NO PASS"
                color = self.green if result == "PASS" else self.red
                self.sheet.cell(row=TEN_ROW, column=5).value = f"{self.convert_data(lt_pkg)}"
                self.sheet.cell(row=TEN_ROW, column=7).fill = color
                self.sheet.cell(row=TEN_ROW, column=7).value = result
                self.tools_result.setdefault("level2 pkg", result)
            elif choice == "L1 and L2 rpm abi":
                lo_abi = similarity.get("level1 rpm abi")
                lt_abi = similarity.get("level2 rpm abi")
                if lo_abi is not None and lt_abi is not None:
                    result = "PASS" if lo_abi >= benchmark["standard"] and \
                                       lt_abi >= benchmark["standard"] else "NO PASS"
                    color = self.green if result == "PASS" else self.red
                    self.sheet.cell(row=12, column=5).value = f"{self.convert_data(lt_abi)}"
                    self.sheet.cell(row=12, column=7).fill = color
                    self.sheet.cell(row=12, column=7).value = result
                    self.tools_result.setdefault("level2 rpm abi", result)
            elif choice == option:
                if data is not None:
                    result = "PASS" if data >= benchmark.get("standard") else "NO PASS"
                    color = self.green if result == "PASS" else self.red
                    self.sheet.cell(row=index, column=5).value = self.convert_data(data)
                    self.sheet.cell(row=index, column=SEVEN_COLUMN).fill = color
                    self.sheet.cell(row=index, column=SEVEN_COLUMN).value = result
                    self.tools_result.setdefault(choice, result)

    def write_summary_file(self, *args):
        """
        write summary file
        Args:
            *args:  similarity, side_a, side_b, root_path, osv_title, iso_path

        Returns:
            None
        """
        try:
            similarity, side_a, side_b, root_path, osv_title, iso_path = args

            self.copy_file(root_path, osv_title)
            benchmark_criteria = self.read_benchmark_json()
            self.open_excel()
            for option, data in similarity.items():
                for index, content in enumerate(benchmark_criteria, start=9):
                    self.result_comparison(similarity, content, option, data, index)
            self.table.save(self.osv_summary_path)
            self.close()
            self.revise_compatible_file(side_a, side_b, iso_path)
        except(AttributeError, KeyError, IOError, OSError, FileNotFoundError, ValueError):
            logger.exception("Excel statistics error")

    def revise_compatible_file(self, side_a, edition, iso_path):
        """
        revise compatible file
        Args:
            side_a: side_a
            edition: edition
            iso_path: iso_path

        Returns:
            None
        """
        self.conclusion = self.read_excel_content()
        os_name = side_a.split('-')[0]
        self.open_excel()
        self.sheet.cell(row=SIX_ROW, column=FIVE_COLUMN).value = SIX_TITLE[0] + os_name + SIX_TITLE[1]
        self.sheet.cell(row=ELEVEN_ROW, column=FOUR_COLUMN).value = ELEVEN_TITLE[0] + os_name + ELEVEN_TITLE[1]
        self.sheet.cell(row=TWELVE_ROW, column=FOUR_COLUMN).value = TWELVE_TITLE[0] + os_name + TWELVE_TITLE[1]
        self.sheet.cell(row=THIRTEEN_ROW, column=FOUR_COLUMN).value = THIRTEEN_TITLE[0] + os_name + THIRTEEN_TITLE[1]
        self.sheet.cell(row=FOURTEEN_ROW, column=FOUR_COLUMN).value = FOURTEEN_TITLE[0] + os_name + FOURTEEN_TITLE[1]
        self.sheet.cell(row=FIFTEEN_ROW, column=FOUR_COLUMN).value = FIFTEEN_TITLE[0] + os_name + FIFTEEN_TITLE[1]
        self.sheet.cell(row=SIX_ROW, column=SIX_COLUMN).value = self.obtain_base_os(side_a)
        self.sheet.cell(row=TWO_ROW, column=SIX_COLUMN).value = edition
        version = "aarch64" if "aarch64" in edition else "X86"
        self.sheet.cell(row=THREE_ROW, column=SIX_COLUMN).value = version
        self.sheet.cell(row=FIVE_ROW, column=SIX_COLUMN).value = gen_hash_key(iso_path)
        font_color = self.green_font if self.conclusion == "通过" else self.red_font
        self.sheet.cell(row=THREE_ROW, column=TWO_COLUMN, value=self.conclusion).font = font_color
        self.table.save(self.osv_summary_path)
        self.close()
