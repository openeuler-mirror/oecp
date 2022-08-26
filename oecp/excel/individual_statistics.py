import logging
import os
import shutil
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd

from oecp.result.constants import *

logger = logging.getLogger("oecp")


class IndividualStatistics:
    """
    Excel table displays statistics for single item comparison information(except for rpm_analyse).
    """

    def __init__(self, report_path):
        self.all_rpm_report = report_path
        self.report_name = "single_calculate.xlsx"
        self.single_calculate_path = None
        self.table = None

    @staticmethod
    def count_by_result(count_content):
        count_row = []
        for cmp_result in COUNT_RESULTS:
            single_count_results = sum(count_content[cmp_result])
            count_row.append(single_count_results)
        return count_row

    @staticmethod
    def count_kernel_analyse(df):
        kernel_result = {}
        for cmp_type in KERNEL_ANALYSE:
            cmp_result = []
            cmp_packages = df.loc[(df[CMP_TYPE] == cmp_type)]
            for count_type in COUNT_RESULTS:
                cmp_result.append(sum(cmp_packages[count_type]))
            kernel_result.setdefault(cmp_type, cmp_result)
        return kernel_result

    @staticmethod
    def insert_abi_diff_rates(df, count_abi_details):
        """
        Insert the abi diff_rates count from file all-rpm-report to rpm abi statistics information.
        Args:
            df: DataFrame.
            count_abi_details(List[List]): all compare type of the rpmabi_analyse.
        Returns:
            count_result(List[List])
        """
        abi_packages = df.loc[(df[CMP_TYPE] == CMP_TYPE_RPM_ABI)]
        abi_packages_diff = abi_packages.loc[abi_packages[CMP_RESULT] == CMP_RESULT_DIFF]
        all_diff_rate = str(len(abi_packages_diff)) + '/' + str(len(abi_packages))
        count_abi_details.get("all_packages_abi").insert(0, all_diff_rate)
        l1_abi_packages = abi_packages.loc[abi_packages[CTG_LEVEL] == 1]
        l1_packsges_diff = l1_abi_packages.loc[l1_abi_packages[CMP_RESULT] == CMP_RESULT_DIFF]
        l1_diff_rate = str(len(l1_packsges_diff)) + '/' + str(len(l1_abi_packages))
        count_abi_details.get("l1_packages_abi").insert(0, l1_diff_rate)
        l2_abi_packages = abi_packages.loc[abi_packages[CTG_LEVEL] == 2]
        l2_packsges_diff = l2_abi_packages.loc[l2_abi_packages[CMP_RESULT] == CMP_RESULT_DIFF]
        l2_diff_rate = str(len(l2_packsges_diff)) + '/' + str(len(l2_abi_packages))
        count_abi_details.get("l2_packages_abi").insert(0, l2_diff_rate)

        return count_abi_details

    def open_report(self):
        """
        Operate the object of single_calculate excel.
        """
        dir_path = Path(Path(__file__).parents[1].joinpath("conf", "excel_template"))
        template_path = Path(dir_path).joinpath(self.report_name)
        self.single_calculate_path = os.path.join(os.path.dirname(self.all_rpm_report), self.report_name)
        shutil.copy(template_path, self.single_calculate_path)

        self.table = load_workbook(self.single_calculate_path)
        sheet_1 = self.table[SHEET_1]
        sheet_2 = self.table[SHEET_2]
        sheet_3 = self.table[SHEET_3]

        return sheet_1, sheet_2, sheet_3

    def write_to_rpmfile(self, count_result, sheet_1):
        """
        Write statistics to rpmfile_analyse.
        Args:
            count_result (List): statistics based on the display information
            sheet_1: sheet of the rpmfile_analyse page
        Returns:
            None
        """
        row_num = 2
        for cmp_result in count_result:
            for single in cmp_result:
                for colu in range(2, 7):
                    sheet_1.cell(row_num, colu, single[colu - 2])
                row_num += 1
            row_num += 2
        self.table.save(self.single_calculate_path)

    def write_to_excel(self, count_result, sheet):
        """
        Write statistics to rpm_analyse/kernel_analyse.
        """
        row = 1
        for single_detail in count_result.values():
            row += 1
            for count_num in range(len(single_detail)):
                sheet.cell(row, count_num + 2, single_detail[count_num])
        self.table.save(self.single_calculate_path)

    def count_by_rpmfile_type(self, df, cmp_types):
        """
        Collected the comparison information in file all-rpm-report by rpm category(all, level 1, level 2).
        Args:
            df: DataFrame.
            cmp_types(List): all compare type of the rpmfile_analyse.
        Returns:
            count_result(List[List])
        """
        count_result = []
        for cmp_type in cmp_types:
            single_result = []
            all_packages = df.loc[(df[CMP_TYPE] == cmp_type)]
            all_packages_diff = all_packages.loc[all_packages[CMP_RESULT] == CMP_RESULT_DIFF]
            all_packages_count = self.count_by_result(all_packages)
            all_packages_change = str(len(all_packages_diff)) + '/' + str(len(all_packages))
            all_packages_count.insert(0, all_packages_change)
            single_result.append(all_packages_count)

            l1_packages = all_packages.loc[all_packages[CTG_LEVEL] == 1]
            l1_packages_diff = l1_packages.loc[l1_packages[CMP_RESULT] == CMP_RESULT_DIFF]
            l1_packages_count = self.count_by_result(l1_packages)
            l1_packages_diff_rate = str(len(l1_packages_diff)) + '/' + str(len(l1_packages))
            l1_packages_count.insert(0, l1_packages_diff_rate)
            single_result.append(l1_packages_count)

            l2_packages = all_packages.loc[all_packages[CTG_LEVEL] == 2]
            l2_packages_diff = l2_packages.loc[l2_packages[CMP_RESULT] == CMP_RESULT_DIFF]
            l2_packages_count = self.count_by_result(l2_packages)
            l2_packages_diff_rate = str(len(l2_packages_diff)) + '/' + str(len(l2_packages))
            l2_packages_count.insert(0, l2_packages_diff_rate)
            single_result.append(l2_packages_count)
            count_result.append(single_result)
        return count_result

    def run_statistics(self, count_abi_details):
        """
        Generate single_calculate excel.
        Args:
            count_abi_details(List): statistical results of abi changes.
        """
        df = pd.read_csv(self.all_rpm_report)
        sheet_1, sheet_2, sheet_3 = self.open_report()
        rpmfile_count = self.count_by_rpmfile_type(df, RPMFILE_CMP_TYPES)
        self.write_to_rpmfile(rpmfile_count, sheet_1)
        count_abi_details = self.insert_abi_diff_rates(df, count_abi_details)
        self.write_to_excel(count_abi_details, sheet_2)
        kernel_result = self.count_kernel_analyse(df)
        self.write_to_excel(kernel_result, sheet_3)
